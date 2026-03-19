"""Export extracted data to Excel (.xlsx)."""

from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
from scipy.interpolate import interp1d

from src.models.project_data import ProjectState
from src.models.series_data import SeriesData
from src.models.types import CombinedMode


def export_to_excel(
    project: ProjectState,
    path: Path,
    combined_mode: Optional[CombinedMode] = None,
) -> None:
    """Write the extracted data to an Excel workbook.

    Creates:
      - One sheet per series (``Series_1``, ``Series_2``, ...)
      - A ``Combined`` sheet
      - A ``Metadata`` sheet
    """
    if combined_mode is None:
        combined_mode = project.combined_mode

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Raw sheets ---
    for sd in project.series:
        sd.sort_by_x()
        ws = wb.create_sheet(title=f"Series_{sd.index}")
        ws.append(["X", "Y"])
        for pt in sd.points:
            ws.append([pt.x, pt.y])

    # --- Combined sheet ---
    _write_combined(wb, project.series, combined_mode)

    # --- Metadata sheet ---
    _write_metadata(wb, project)

    wb.save(str(path))


# -----------------------------------------------------------------------

def _write_combined(
    wb: openpyxl.Workbook,
    series_list: list[SeriesData],
    mode: CombinedMode,
) -> None:
    ws = wb.create_sheet(title="Combined")
    if not series_list:
        return

    all_xs = _build_combined_x(series_list, mode)
    headers = ["X"] + [f"Y_{s.index}" for s in series_list]
    ws.append(headers)

    interps = []
    for sd in series_list:
        xs = np.array(sd.xs)
        ys = np.array(sd.ys)
        if len(xs) >= 2:
            interps.append(interp1d(xs, ys, kind="linear", bounds_error=False, fill_value=np.nan))
        elif len(xs) == 1:
            interps.append(lambda _x, _y=ys[0]: _y)
        else:
            interps.append(lambda _x: np.nan)

    for xi in all_xs:
        row: list[object] = [float(xi)]
        for fn in interps:
            val = float(fn(xi))
            row.append(val if not math.isnan(val) else None)
        ws.append(row)


def _build_combined_x(series_list: list[SeriesData], mode: CombinedMode) -> list[float]:
    all_x_flat = []
    for sd in series_list:
        all_x_flat.extend(sd.xs)

    if not all_x_flat:
        return []

    if mode == CombinedMode.UNION_X:
        return sorted(set(all_x_flat))

    x_min, x_max = min(all_x_flat), max(all_x_flat)
    if mode == CombinedMode.UNIFORM_GRID:
        n = max(len(all_x_flat), 200)
        return list(np.linspace(x_min, x_max, n))

    # INTERPOLATION: use the densest series' grid
    n = max(max(len(sd.xs) for sd in series_list), 200)
    return list(np.linspace(x_min, x_max, n))


def _write_metadata(wb: openpyxl.Workbook, project: ProjectState) -> None:
    ws = wb.create_sheet(title="Metadata")
    rows = [
        ("Parameter", "Value"),
        ("source_file", str(project.image_path or "")),
        ("date", datetime.now().isoformat(timespec="seconds")),
        ("x_scale", project.settings.x_scale.name),
        ("y_scale", project.settings.y_scale.name),
        ("series_count", len(project.series)),
    ]
    for sd in project.series:
        rows.append((f"series_{sd.index}_mode", sd.mode.name))
        rows.append((f"series_{sd.index}_kind", sd.kind.name))
        rows.append((f"series_{sd.index}_points", len(sd.points)))
    for r in rows:
        ws.append(list(r))
